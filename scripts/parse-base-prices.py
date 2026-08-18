#!/usr/bin/env python3
"""Parse OneDrive Base Price List 2025.xlsx → lib/basePrices.data.json.

Never commit the xlsx. Safe to re-run whenever the list is updated.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("need openpyxl (use /tmp/oltc-extract or pip install openpyxl)")

ONEDRIVE = (
    Path.home()
    / "Library/CloudStorage/OneDrive-上海华明电力设备制造有限公司"
    / "QS"
    / "a. Base Price List 2025.xlsx"
)
OUT = Path(__file__).resolve().parents[1] / "lib" / "basePrices.data.json"

# Sheets we care about for the public selector (standard OLTC sets).
# Skip extras / OCTC / MDU-only.
KEEP_SHEETS = {
    "CV & SV",
    "CV2",
    "CM2III",
    "CM2 II",
    "CM2 I",
    "CMIII",
    "CMII",
    "CMI",
    "SHZVIII",
    "SHZVII",
    "SHZVI",
    "SHZVG III",
    "SHZVG II",
    "SHZVG I",
    "CZ & CVT",
    "CMDIII",
    "CMDII",
    "CMDI",
    "HWV",
    "HWDK",
    "WSL(WDL)",
}

# Longest family first so CV2/CM2/SHZVG/CVIII do not split as CV+II / CVI+I.
FAM = r"(?P<family>SHZVG|SHZV|HWDK|HWV|CMD|CV2|CM2|CVT|CV|SV|CM|CZ)"
OCTC_ROMAN = r"(?P<series>VIII|VII|III|II|IV|VI|IX|V|I)"
# Known Um values, longest first, so 1268x7 → 126-8x7
OCTC_UMS = (
    "363",
    "362",
    "330",
    "300",
    "252",
    "170",
    "145",
    "126",
    "72.5",
    "40.5",
    "17.5",
    "12.5",
    "12",
)
PHASE = r"(?P<phases>III|II|I)"
TAIL = (
    r"-"
    r"(?P<current>\d+)"
    r"(?P<conn>[YD])?"
    r"/"
    r"(?P<um>\d+(?:\.\d+)?)"
    r"(?P<grade>DE|B|C|D|E)?"
)

# Family, phases, current, optional Y/D, Um, optional grade, pitch dots, optional W/G/0
ROW_RE = re.compile(
    rf"^{FAM}{PHASE}{TAIL}"
    r"(?:-"
    r"(?P<pitch>\d{2})"
    r"(?:\.{2,3}|…)"
    r"(?P<reg>[WG0])?"
    r")?"
    r"$",
    re.I,
)

# HWV / HWDK list rows carry a full tap code (usually 18353W)
FULL_RE = re.compile(
    rf"^{FAM}{PHASE}{TAIL}"
    r"-"
    r"(?P<tap>\d{5}[WG0]?)$",
    re.I,
)

# CZ list is 3× singles with a linear position count, not pitch-dots.
CZ_RE = re.compile(
    r"^(?:(?P<units>\d+)[x×])?CZ(?P<phases>III|II|I)-"
    r"(?P<current>\d+)/"
    r"(?P<um>\d+(?:\.\d+)?)-"
    r"(?P<pos>\d+)$",
    re.I,
)

# CVT list uses a bare position count (7/9/13).
CVT_RE = re.compile(
    r"^CVT(?P<phases>III|II|I)-"
    r"(?P<current>\d+)"
    r"(?P<conn>[YD])?/"
    r"(?P<um>\d+(?:\.\d+)?)-"
    r"(?P<pos>\d+)$",
    re.I,
)


def cells(row) -> list:
    return [c.value for c in row]


def first_str(vals) -> str | None:
    for v in vals:
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def first_num(vals) -> float | None:
    for v in vals:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if v > 0:
                return float(v)
    return None


def compact_octc_label(raw: str) -> str:
    s = raw.replace("\xa0", " ")
    s = re.sub(r"\s+", "", s)
    s = s.replace("×", "x").replace("*", "x").replace("X", "x")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("－", "-").replace("–", "-").replace("—", "-")
    s = s.replace("／", "/")
    s = s.replace(",", ".")
    s = re.sub(r"(\d)_(\d)", r"\1x\2", s)
    s = re.sub(r"(III|II|I)(\d)", r"\1-\2", s, count=1)
    s = re.sub(r"(WSL|WDL|WSG)(VIII|VII|IV|VI|IX|V)(\d)", r"\1\2-\3", s, count=1, flags=re.I)
    if "/" not in s:
        s = re.sub(r"([YD])(\d+(?:\.\d+)?)-", r"\1/\2-", s, count=1, flags=re.I)
    um_alt = "|".join(re.escape(u) for u in OCTC_UMS)
    s = re.sub(rf"({um_alt})(\d{{1,2}}x\d)", r"\1-\2", s, count=1)
    return s


def parse_wsl_label(raw: str) -> dict | None:
    """WSL/WDL commercial string → list row. Contact is last NxM (4x3(6x5)→6x5)."""
    s = compact_octc_label(raw)
    s = re.sub(
        r"(withhandwheel|withmanualdrive|手动|顶盖.*|头部.*|钟|串并联.*).*$",
        "",
        s,
        flags=re.I,
    )
    m = re.match(
        rf"^(?P<family>WSL|WDL|WSG){OCTC_ROMAN}-"
        r"(?P<current>\d+)"
        r"(?P<conn>[YD])?"
        r"[/-]"
        r"(?P<um>\d+(?:\.\d+)?)"
        r"-"
        r"(?P<tail>.+)$",
        s,
        re.I,
    )
    if not m:
        return None
    tail = m.group("tail")
    pairs = re.findall(r"(\d+)\s*x\s*(\d+)", tail, re.I)
    if not pairs:
        return None
    contact = f"{int(pairs[-1][0])}x{int(pairs[-1][1])}"
    size_m = re.search(r"\(([A-E])\)\s*$|([A-E])\s*$", tail, re.I)
    size = ""
    if size_m:
        size = (size_m.group(1) or size_m.group(2) or "").upper()
    um = float(m.group("um"))
    family = m.group("family").upper()
    series = m.group("series").upper()
    conn = (m.group("conn") or "").upper()
    current = int(m.group("current"))
    list_key = f"{family}{series}-{current}{conn}/{um:g}-{contact}{size}"
    return {
        "family": family,
        "phases": series,
        "currentA": current,
        "connection": conn,
        "umKv": um,
        "selectorSize": size,
        "pitch": None,
        "changeOver": None,
        "listKey": list_key,
        "tapCode": contact,
        "unitCount": 1,
    }


def parse_label(raw: str) -> dict | None:
    s = re.sub(r"\s+", "", raw)
    s = s.replace("×", "x")
    s = s.replace("…", "...").replace("..", "...")
    s = re.sub(r"\.{2,}", "...", s)
    m = CZ_RE.match(s)
    if m:
        d = m.groupdict()
        return {
            "family": "CZ",
            "phases": d["phases"].upper(),
            "currentA": int(d["current"]),
            "connection": "",
            "umKv": float(d["um"]),
            "selectorSize": "",
            "pitch": None,
            "changeOver": "0",
            "listKey": s,
            "tapCode": d["pos"],
            "unitCount": int(d["units"] or 1),
        }
    m = CVT_RE.match(s)
    if m:
        d = m.groupdict()
        return {
            "family": "CVT",
            "phases": d["phases"].upper(),
            "currentA": int(d["current"]),
            "connection": (d.get("conn") or "").upper(),
            "umKv": float(d["um"]),
            "selectorSize": "",
            "pitch": None,
            "changeOver": "0",
            "listKey": s,
            "tapCode": d["pos"],
            "unitCount": 1,
        }
    m = FULL_RE.match(s) or ROW_RE.match(s)
    if not m:
        return None
    d = m.groupdict()
    family = d["family"].upper()
    pitch = None
    change_over = None
    tap = d.get("tap")
    if tap:
        pitch = int(tap[:2])
        last = tap[-1].upper()
        # 10070 / 18353W / missing letter: linear is 0, never leave null
        change_over = last if last in "WG0" else "0"
    else:
        if d.get("pitch"):
            pitch = int(d["pitch"])
        reg = (d.get("reg") or "").upper()
        if reg in ("W", "G", "0"):
            change_over = reg
        elif pitch is not None:
            # CV2 writes linear as 10... (no trailing 0)
            change_over = "0"
        else:
            change_over = None
    return {
        "family": family,
        "phases": d["phases"].upper(),
        "currentA": int(d["current"]),
        "connection": (d.get("conn") or "").upper(),
        "umKv": float(d["um"]),
        "selectorSize": (d.get("grade") or "").upper(),
        "pitch": pitch,
        "changeOver": change_over,
        "listKey": s,
        "tapCode": tap or "",
        "unitCount": 1,
    }


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else ONEDRIVE
    if not src.exists():
        sys.exit(f"missing price list: {src}")

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    rows: list[dict] = []
    skipped_sheets: list[str] = []
    unread_labels = 0

    for name in wb.sheetnames:
        if name not in KEEP_SHEETS:
            skipped_sheets.append(name)
            continue
        ws = wb[name]
        wsl_sheet = name == "WSL(WDL)"
        for row in ws.iter_rows(max_col=8):
            vals = cells(row)
            label = first_str(vals)
            if not label:
                continue
            parsed = parse_wsl_label(label) if wsl_sheet else parse_label(label)
            if not parsed:
                continue
            if wsl_sheet:
                # Col E = CMA7 motor set (hand wheel / manual / SHM-D / CMA7 / remote).
                # Skip rows with no CMA7 — do not invent from hand-wheel.
                cma7 = vals[4] if len(vals) > 4 else None
                price = (
                    float(cma7)
                    if isinstance(cma7, (int, float)) and not isinstance(cma7, bool) and cma7 > 0
                    else None
                )
            else:
                price = first_num(vals[1:])  # never treat the model cell as the price
            if price is None:
                unread_labels += 1
                continue
            parsed["listRmb"] = int(round(price))
            parsed["sheet"] = name
            parsed["includesMdu"] = True  # standard set; HWV / WSL CMA7 already includes MDU
            rows.append(parsed)
    wb.close()

    # Dedupe exact listKey, keep first
    seen: set[str] = set()
    unique: list[dict] = []
    for r in rows:
        k = r["listKey"]
        if k in seen:
            continue
        seen.add(k)
        unique.append(r)

    payload = {
        "source": "QS/a. Base Price List 2025.xlsx",
        "generatedOn": date.today().isoformat(),
        "currency": "CNY",
        "incoterm": "FOB Shanghai",
        "includesMdu": True,
        "rowCount": len(unique),
        "skippedSheets": skipped_sheets,
        "headerRowsWithoutPrice": unread_labels,
        "rows": unique,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    print(f"wrote {OUT} ({len(unique)} rows, {unread_labels} headers, skipped {skipped_sheets})")


if __name__ == "__main__":
    main()
