#!/usr/bin/env python3
"""Extract sold OLTC type + transformer duty from 2026 OS/PO PDFs.

Reads first 12 pages with pdfplumber (no OCR). Writes Excel.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import traceback
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

PDF_ROOT = Path(r"C:\Users\TYM\Desktop\2026")
OUT_XLSX = Path(r"C:\Users\TYM\projects\oltc-selector\docs\2026-os-sales.xlsx")
OUT_COPY = Path(r"C:\Users\TYM\Desktop\2026-os-sales.xlsx")

INCLUDE_FOLDERS = {
    "CM", "CMD", "CV", "CVT", "CZ", "HWV", "SHZV", "SY", "VCM", "VCV", "WG", "WL",
}
SKIP_FOLDERS = {"MDU", "ZXJY", "售后芯子+油室+散件"}
FOLDER_ALIAS = {"VCM": "CM2", "VCV": "CV2", "WL": "WSL", "WG": "WSG"}

# Longest-first so CM2/CV2/CMD/SHZVG win over CM/CV/SHZV.
FAMILIES = (
    "SHZVG", "SHZV", "HWDK", "HWV", "CMD", "CM2", "CVT", "CV2",
    "WSL", "WDL", "WSG", "WDG", "WLG", "WLL", "SYXZZ", "SYXZ", "SY",
    "SV", "CZ", "CM", "CV",
)
FAM = "|".join(FAMILIES)
ROMAN = r"VIII|VII|III|II|IV|VI|IX|V|I"
PHASE_TOKEN = rf"(?:{ROMAN}|3|1|111)"

MAX_PAGES = 12
PER_FILE_S = 20
DEFAULT_WORKERS = 6

SALES_COLS = [
    "file", "folder", "serial", "customer", "family_folder",
    "sold_type", "type_alts", "phases", "current_a", "connection",
    "um_kv", "selector_size", "tap_code", "unit_count",
    "mva", "rated_kv", "i_a", "i_max_a", "ust_v", "ust_max_v",
    "plus_minus_steps", "plus_steps", "minus_steps",
    "mdu", "text_status", "notes",
]

CYR_LATIN = str.maketrans({
    "А": "A", "В": "B", "С": "C", "Е": "E", "Н": "H", "К": "K", "М": "M",
    "О": "O", "Р": "P", "Т": "T", "Х": "X", "У": "Y",
    "а": "a", "с": "c", "е": "e", "о": "o", "р": "p", "х": "x", "у": "y",
    "З": "3", "з": "3", "І": "I", "і": "i",
    "\u00a0": " ",
})

DATA_MARKERS = re.compile(
    r"(ON[- ]LOAD TAP CHANGER\s*DATA|OFF[- ]CIRCUIT TAP CHANGER\s*DATA|"
    r"TAP CHANGER\s*DATA|TapChanger Data|Tap Changer Data|"
    r"TYPE CZ TAP CHANGER\s*DATA|TYPE CVT TAP CHANGER\s*DATA|"
    r"ON-LOAD TAP GHANGER\s*DATA|ДАННЫЕ УСТРОЙСТВА)",
    re.I,
)
XFMR_MARKERS = re.compile(
    r"(TRANSFORMER DATA|Transformer Data|Rated capacity|Rated power|"
    r"ДАННЫЕ ТРАНСФОРМАТОРА|Ном\.\s*мощность)",
    re.I,
)

# Compact: CM2III-500Y/72.5C-12233G  /  3xCZI-500/40.5-7
COMPACT_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?P<fam>{FAM})"
    rf"\s*-?\s*(?P<ph>{ROMAN}|I)"
    rf"\s*-?\s*"
    rf"(?P<cur>\d{{2,5}})"
    rf"\s*(?P<yd>[YDyd])?"
    rf"\s*/\s*"
    rf"(?P<um>\d+(?:[.,]\d+)?)"
    rf"\s*(?P<sel>[BCDE]{{1,2}})?"
    rf"\s*-\s*"
    rf"(?P<tap>\d{{2}}\.\d{{2}}\.\d[WGwg0]?|\d{{4,5}}\s*[WGwg0]?|\d{{1,2}}x\d(?:\(\d+x\d\))?[A-Ea-e]?|\d{{1,2}})",
)

# Filename / no-slash: CV2III350D-40.5-10193W  or  CM2III-500Y-170/B-14273G
FILEISH_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?P<fam>{FAM})"
    rf"\s*-?\s*(?P<ph>{ROMAN})"
    rf"\s*-?\s*"
    rf"(?P<cur>\d{{2,5}})"
    rf"\s*(?P<yd>[YDyd])"
    rf"\s*[-/]\s*"
    rf"(?P<um>\d+(?:[.,]\d+)?)"
    rf"\s*[/]?\s*(?P<sel>[BCDE]{{1,2}})?"
    rf"\s*-?\s*"
    rf"(?P<tap>\d{{2}}\.\d{{2}}\.\d[WGwg0]?|\d{{4,5}}\s*[WGwg0]?)",
)

# Spaced OS table / PO:
#   CM III 600 Y 72,5 C 14253W
#   1x CV III 350 D 40.5 12233G
#   HWV 3 400 Y 17.5 18353W
#   WSL IV 600 D 72.5 6x5(B)
#   CV2 III 350 D 145-10193W
OS_ROW_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?:COM\.)?(?P<fam>{FAM})"
    rf"\s+(?P<ph>{PHASE_TOKEN})"
    rf"\s+(?P<cur>\d{{2,5}})"
    rf"(?:\s+(?P<yd>[YDyd]))?"
    rf"\s+(?:(?:kV\s*)?(?P<um>\d+(?:[.,]\s*\d+)?)\s*(?:kV)?)"
    rf"(?:\s*(?:[/]\s*)?(?P<sel>[BCDE]{{1,2}}))?"
    rf"\s*-?\s*"
    rf"(?P<tap>\d{{2}}\.\d{{2}}\.\d[WGwg0]?|\d{{4,5}}[ \t]*[WGwg0]?|\d+[ \t]*[xX][ \t]*\d+(?:[ \t]*\([A-Ea-e]\))?[A-Ea-e]?|\d{{3,4}})",
)

# Wilson: SHZV III 1000 Y-126/D 10.19.3W
# GETRA:  CM2 III 500Y-170/B 14273G
WILSON_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?:COM\.)?(?P<fam>{FAM})"
    rf"\s+(?P<ph>{PHASE_TOKEN})"
    rf"\s+(?P<cur>\d{{2,5}})"
    rf"\s*(?P<yd>[YDyd])"
    rf"\s*[-/]\s*"
    rf"(?P<um>\d+(?:[.,]\d+)?)"
    rf"\s*[/]?\s*(?P<sel>[BCDE]{{1,2}})?"
    rf"\s+"
    rf"(?P<tap>\d{{2}}\.\d{{2}}\.\d[WGwg0]?|\d{{4,5}}\s*[WGwg0]?)",
)

# HWV PO: HWV III 400 Y 40.5 kV 18353W
HWV_PO_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?P<fam>HWV|HWDK)"
    rf"\s+(?P<ph>{PHASE_TOKEN})"
    rf"\s+(?P<cur>\d{{2,5}})"
    rf"\s+(?P<yd>[YDyd])"
    rf"\s+(?P<um>\d+(?:[.,]\d+)?)\s*kV"
    rf"\s+(?P<tap>\d{{4,5}}\s*[WGwg0]?)",
)

# SY: SYXZZ-40.5-200-9
SY_RE = re.compile(
    r"(?i)(?P<fam>SYXZZ|SYXZ|SYX)\s*-?\s*"
    r"(?P<um>\d+(?:[.,]\d+)?)\s*-?\s*"
    r"(?P<cur>\d{2,4})\s*-?\s*"
    r"(?P<tap>\d{1,2})\b",
)

# OCTC compact: WSLIV-800Y/170-6x5B
OCTC_COMPACT_RE = re.compile(
    rf"(?i)(?:(?P<n>\d+)\s*[xX×]\s*)?"
    rf"(?P<fam>WSL|WDL|WSG|WDG|WLG|WLL)"
    rf"\s*-?\s*(?P<ph>{ROMAN})"
    rf"\s*-?\s*(?P<cur>\d{{2,5}})"
    rf"\s*(?P<yd>[YDyd])?"
    rf"\s*[/-]\s*(?P<um>\d+(?:[.,]\d+)?)"
    rf"\s*-?\s*(?P<tap>\d+\s*[xX×*]\s*\d+(?:\s*\(\d+\s*[xX×]\s*\d+\))?)"
    rf"\s*(?:\((?P<sel1>[A-Ea-e])\)|(?P<sel2>[A-Ea-e]))?",
)

MDU_ASSIGN_RE = re.compile(
    r"(?i)(?:motor\s*drive(?:\s*unit|\s*type)?|MDU|\+)\s*[:.]?\s*"
    r"(CMA\s*7|SHM-?DL|SHM-?D|SHM-?K|SHM-?KX|HMC-3C|ET-SZ6|HMJK-?10Z?|HMBK-35)",
)
MDU_TOKEN_RE = re.compile(
    r"(?i)\b(CMA\s*7|SHM-?DL|SHM-?D|SHM-?K|SHM-?KX|HMC-3C|ET-SZ6|HMJK-?10Z?|HMBK-35)\b",
)
RANGE_MDU_LINE = re.compile(
    r"(?i)CMA\s*7\s+SHM-?D|Motor Drive Unit|Моторный привод",
)

SERIAL_RE = re.compile(
    r"^(E-[A-Z]*\d{5,8}(?:\s*[-,]\s*(?:E-[A-Z]*)?\d+)*)",
    re.I,
)


def fold_text(s: str) -> str:
    if not s:
        return ""
    s = s.translate(CYR_LATIN)
    s = (
        s.replace("，", ",")
        .replace("．", ".")
        .replace("＝", "=")
        .replace("（", "(")
        .replace("）", ")")
        .replace("－", "-")
        .replace("–", "-")
        .replace("—", "-")
        .replace("×", "x")
        .replace("√", "/")
        .replace("±", "±")
        .replace("cid:713", "-")
        .replace("cid:730", ">")
    )
    s = re.sub(r"[ \t]+", " ", s)
    return s


def parse_num(raw: str | None) -> float | None:
    if not raw:
        return None
    s = fold_text(str(raw)).strip()
    if "=" in s:
        s = s.split("=")[-1]
    s = s.replace(" ", "").replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def num_or_none(v: float | None) -> float | int | None:
    if v is None:
        return None
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return v


def norm_tap(raw: str | None) -> str:
    if not raw:
        return ""
    t = fold_text(raw)
    t = re.sub(r"\s+", "", t).upper()
    t = t.replace("*", "x").replace("X", "x")
    # after lowercasing x, restore W/G already upper; digits stay
    t = t.replace("w", "W").replace("g", "G")
    m = re.fullmatch(r"(\d{2})\.(\d{2})\.(\d)([WG0]?)", t)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}{m.group(4)}"
    m = re.fullmatch(r"(\d+)\s*X\s*(\d+)(?:\((\d+)X(\d+)\))?(?:\(([A-E])\)|([A-E]))?", t)
    if m:
        inner = f"{m.group(3)}x{m.group(4)}" if m.group(3) else f"{m.group(1)}x{m.group(2)}"
        return inner
    t = re.sub(r"(\d)\s+([WG])$", r"\1\2", t)
    return t


def norm_phases(raw: str | None) -> str:
    if not raw:
        return ""
    p = fold_text(raw).upper().replace(" ", "")
    if p in {"3", "111"}:
        return "III"
    if p == "1":
        return "I"
    return p


def norm_family(raw: str | None) -> str:
    if not raw:
        return ""
    f = fold_text(raw).upper().replace(" ", "")
    if f in {"SYX", "SYXZ"}:
        return "SYXZZ"
    if f == "CVT" or f == "CVТ":
        return "CVT"
    return f


def um_str(um: float | None) -> str:
    if um is None:
        return ""
    if abs(um - round(um)) < 1e-9:
        return str(int(round(um)))
    s = f"{um:.4f}".rstrip("0").rstrip(".")
    return s


def format_sold_type(p: dict) -> str | None:
    fam = p.get("family") or ""
    ph = p.get("phases") or ""
    cur = p.get("current_a")
    conn = (p.get("connection") or "").upper()
    um = p.get("um_kv")
    sel = (p.get("selector_size") or "").upper()
    tap = p.get("tap_code") or ""
    n = p.get("unit_count") or 1
    if not fam or um is None or cur is None:
        return None
    ums = um_str(um)
    cur_s = str(int(cur)) if float(cur) == int(cur) else str(cur)
    if fam in {"WSL", "WDL", "WSG", "WDG", "WLG", "WLL"}:
        core = f"{fam}{ph}-{cur_s}{conn}/{ums}-{tap}{sel}"
    elif fam.startswith("SY"):
        core = f"{fam}-{ums}-{cur_s}-{tap}" if tap else f"{fam}-{ums}-{cur_s}"
    elif fam == "CZ":
        core = f"{fam}{ph}-{cur_s}/{ums}-{tap}" if tap else f"{fam}{ph}-{cur_s}/{ums}"
    elif fam in {"CV", "CV2", "SV", "CVT", "HWV", "HWDK"}:
        core = f"{fam}{ph}-{cur_s}{conn}/{ums}-{tap}" if tap else f"{fam}{ph}-{cur_s}{conn}/{ums}"
    else:
        token = f"{ums}{sel}" if sel else ums
        if conn:
            core = f"{fam}{ph}-{cur_s}{conn}/{token}-{tap}" if tap else f"{fam}{ph}-{cur_s}{conn}/{token}"
        else:
            core = f"{fam}{ph}-{cur_s}/{token}-{tap}" if tap else f"{fam}{ph}-{cur_s}/{token}"
    try:
        n_i = int(n)
    except (TypeError, ValueError):
        n_i = 1
    if n_i > 1:
        return f"{n_i}x{core}"
    return core


def hit_from_match(m: re.Match, source: str) -> dict | None:
    gd = m.groupdict()
    fam = norm_family(gd.get("fam"))
    if not fam:
        return None
    cur = parse_num(gd.get("cur"))
    um = parse_num((gd.get("um") or "").replace(" ", ""))
    if cur is None or um is None:
        return None
    if cur < 20 or cur > 8000:
        return None
    if um < 3 or um > 800:
        return None
    sel = (gd.get("sel") or gd.get("sel1") or gd.get("sel2") or "").upper()
    tap_raw = gd.get("tap") or ""
    tap = norm_tap(tap_raw)
    if sel and tap.endswith(sel) and "x" in tap.lower():
        tap = tap[: -len(sel)]
    # 6x5(B) → selector B already captured
    m_oct = re.fullmatch(r"(\d+X\d+)\(([A-E])\)", fold_text(tap_raw).upper().replace(" ", ""))
    if m_oct:
        tap = m_oct.group(1).replace("X", "x")
        sel = sel or m_oct.group(2)
    n = parse_num(gd.get("n"))
    unit = int(n) if n and n > 1 else 1
    yd = (gd.get("yd") or "").upper()
    # CV/CV2/SV/HWV/CVT: a lone D/Y next to current is connection, not selector.
    compound = fam in {"CV", "CV2", "SV", "CVT", "HWV", "HWDK", "CZ"}
    if compound and sel in {"Y", "D"} and not yd:
        yd, sel = sel, ""
    ph = norm_phases(gd.get("ph"))
    hit = {
        "family": fam,
        "phases": ph,
        "current_a": num_or_none(cur),
        "connection": yd,
        "um_kv": num_or_none(um),
        "selector_size": sel,
        "tap_code": tap,
        "unit_count": unit,
        "source": source,
        "raw": m.group(0).strip(),
    }
    hit["sold_type"] = format_sold_type(hit)
    return hit


def zone_after(text: str, rx: re.Pattern, width: int = 1800) -> str:
    chunks = []
    for m in rx.finditer(text):
        chunks.append(text[m.start() : m.start() + width])
    return "\n".join(chunks)


def find_types(text: str, family_hint: str = "") -> list[dict]:
    hits: list[dict] = []
    oltc_zone = zone_after(text, DATA_MARKERS, 1600)
    blobs = [("oltc_data", oltc_zone), ("body", text)]
    specs = [
        ("compact", COMPACT_RE),
        ("fileish", FILEISH_RE),
        ("wilson", WILSON_RE),
        ("hwv_po", HWV_PO_RE),
        ("os_row", OS_ROW_RE),
        ("octc", OCTC_COMPACT_RE),
        ("sy", SY_RE),
    ]
    seen_raw = set()
    for source_blob, blob in blobs:
        if not blob:
            continue
        for source, rx in specs:
            for m in rx.finditer(blob):
                raw = m.group(0).strip()
                key = re.sub(r"\s+", "", raw.upper())
                if key in seen_raw:
                    continue
                # RANGE checkbox line is concatenated families, not a type row.
                if re.search(r"(CMIII|CMDIII|CVIII|CM2III|CV2III|SHZVIII).{0,8}(CMI|CVI|CMD)", raw):
                    continue
                hit = hit_from_match(m, source_blob if source_blob == "oltc_data" else source)
                if not hit or not hit.get("sold_type"):
                    continue
                seen_raw.add(key)
                hits.append(hit)
    return hits


def family_matches_hint(fam: str, hint: str) -> bool:
    if not fam or not hint:
        return False
    if fam == hint:
        return True
    if hint == "SY" and fam.startswith("SY"):
        return True
    if hint == "WSG" and fam in {"WSG", "WDG", "WLG"}:
        return True
    if hint == "WSL" and fam in {"WSL", "WDL", "WLL"}:
        return True
    return False


def score_hit(h: dict, family_hint: str) -> tuple:
    fam = h.get("family") or ""
    hint_ok = 1 if family_matches_hint(fam, family_hint) else 0
    src = h.get("source") or ""
    src_rank = 0 if src == "oltc_data" else 1 if src in {
        "compact", "os_row", "octc", "fileish", "wilson", "hwv_po", "sy",
    } else 2
    tap_ok = 1 if h.get("tap_code") else 0
    sel_ok = 1 if h.get("selector_size") else 0
    length = len(h.get("sold_type") or "")
    return (-hint_ok, src_rank, -tap_ok, -sel_ok, -length)


def pick_type(hits: list[dict], family_hint: str) -> tuple[dict | None, str]:
    if not hits:
        return None, ""
    ranked = sorted(hits, key=lambda h: score_hit(h, family_hint))
    best = ranked[0]
    alts = []
    seen = {best.get("sold_type")}
    for h in ranked[1:]:
        st = h.get("sold_type")
        if st and st not in seen:
            seen.add(st)
            alts.append(st)
    return best, " | ".join(alts[:8])


def parse_serial_customer(stem: str) -> tuple[str, str]:
    m = SERIAL_RE.match(stem.strip())
    if not m:
        return stem.strip(), ""
    serial = re.sub(r"\s+", "", m.group(1))
    rest = stem[m.end() :]
    rest = rest.lstrip(" -_,.")
    return serial, rest.strip()


def range_unit_count(serial: str) -> int | None:
    if not serial:
        return None
    parts = re.findall(r"E-[A-Z]*\d{5,8}", serial, re.I)
    if len(parts) > 1:
        return len(parts)
    m = re.search(r"(\d{5,8})\s*-\s*(\d{1,8})$", serial)
    if not m:
        return None
    a = int(m.group(1))
    b_s = m.group(2)
    b = int(b_s)
    if b < a and len(b_s) < len(str(a)):
        suf = a % (10 ** len(b_s))
        if b >= suf:
            return b - suf + 1
        return None
    if b >= a:
        n = b - a + 1
        return n if 1 < n <= 80 else None
    return None


def first_num_in(text: str, rx: re.Pattern) -> float | None:
    m = rx.search(text)
    if not m:
        return None
    return parse_num(m.group(1))


def extract_transformer(text: str) -> dict:
    zone = zone_after(text, XFMR_MARKERS, 2200) or text
    # Avoid OLTC Um / insulation table leaking into transformer kV.
    xfmr = zone
    out: dict = {}

    mva_vals = []
    for rx in (
        re.compile(r"\(\s*(\d+(?:[.,]\d+)?)\s*\)\s*MVA", re.I),
        re.compile(r"(?:Rated\s*(?:capacity|power)|Ном\.\s*мощность)[^\n]{0,90}?(\d+(?:[.,]\d+)?)\s*MVA", re.I),
        re.compile(r"(\d+(?:[.,]\d+)?)\s*MVA", re.I),
    ):
        for m in rx.finditer(xfmr):
            v = parse_num(m.group(1))
            if v is not None and 0.05 <= v <= 2000:
                mva_vals.append(v)
        if mva_vals:
            break
    if not mva_vals:
        kva_hits = re.findall(
            r"(\d+(?:[.,]\d+)?)(?:\s*/\s*(\d+(?:[.,]\d+)?))?\s*kVA",
            xfmr,
            re.I,
        )
        for a, b in kva_hits:
            v = parse_num(b or a)
            if v is not None and 20 <= v <= 2_000_000:
                mva_vals.append(v / 1000.0)
    if mva_vals:
        out["mva"] = num_or_none(mva_vals[0])

    kv_vals = []
    for rx in (
        re.compile(
            r"(?:regulated voltage|regulating voltage|регулируемое[^\n]{0,30})\s*"
            r"(\d+(?:[.,]\d+)?)\s*\)?\s*kV",
            re.I,
        ),
        re.compile(
            r"\(\s*(\d+(?:[.,]\d+)?)\s*[-–]\s*(?:regulating|regulated)\s*voltage\s*\)\s*kV",
            re.I,
        ),
        re.compile(r"High\s*Voltage\s*(\d+(?:[.,]\d+)?)\s*kV", re.I),
        re.compile(r"\(\s*(\d+(?:[.,]\d+)?)\s*\)\s*kV", re.I),
        re.compile(
            r"(?:Rated\s*voltage|Ном\.\s*напряжение|Regulated\s*voltage)[^\n]{0,80}?"
            r"(\d+(?:[.,]\d+)?)\s*kV",
            re.I,
        ),
    ):
        for m in rx.finditer(xfmr):
            v = parse_num(m.group(1))
            if v is not None and 1 <= v <= 800:
                kv_vals.append(v)
        if kv_vals:
            break
    if kv_vals:
        out["rated_kv"] = num_or_none(kv_vals[0])

    def curr_from(label_rx: str, field: str) -> None:
        rx = re.compile(label_rx, re.I)
        m = rx.search(xfmr)
        if not m:
            return
        blob = m.group(1)
        # I=(515/√3 = 297) already folded
        if "=" in blob:
            v = parse_num(blob)
        else:
            parts = [parse_num(p) for p in re.split(r"[/]", blob)]
            parts = [p for p in parts if p is not None]
            v = parts[-1] if field == "i_max_a" and parts else (parts[0] if parts else None)
        if v is not None and 5 <= v <= 8000:
            out[field] = num_or_none(v)

    curr_from(r"(?<![A-Za-z])I\s*=\s*\(\s*([^)]{1,40})\s*\)\s*A", "i_a")
    curr_from(r"(?<![A-Za-z])I\s*=\s*\(\s*([^)]{1,40})\s*\)\s*[AАд]", "i_a")
    curr_from(r"(?<![A-Za-z])Imax\.?\s*=\s*\(\s*([^)]{1,40})\s*\)\s*A", "i_max_a")
    curr_from(r"\blmax\.?\s*=\s*\(\s*([^)]{1,40})\s*\)\s*A", "i_max_a")
    if "i_a" not in out:
        curr_from(r"Nominal\s*current\s*([0-9,.\s/]{1,40}?)\s*A", "i_a")
    if "i_max_a" not in out:
        curr_from(r"Max current\s*([0-9,.\s/]{1,40}?)\s*A", "i_max_a")
    if "i_a" not in out:
        curr_from(r"Transformer rated\s*current[^\n]{0,20}I\s*=\s*\(?\s*([^)\n]{1,40}?)\s*\)?\s*A", "i_a")

    ust = first_num_in(xfmr, re.compile(r"Ust\s*=\s*\(\s*(\d+(?:[.,]\d+)?)\s*\)\s*V", re.I))
    ust_max = first_num_in(xfmr, re.compile(r"Ust\s*max\.?\s*=\s*\(\s*(\d+(?:[.,]\d+)?)\s*\)\s*V", re.I))
    if ust is None:
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*V\s*[,，]?\s*Constant", xfmr, re.I)
        if m:
            ust = parse_num(m.group(1))
    if ust_max is None:
        m = re.search(r"Max\s*=\s*(\d+(?:[.,]\d+)?)\s*V", xfmr, re.I)
        if m:
            ust_max = parse_num(m.group(1))
            # WEG form sometimes writes winding voltage (144900 V) as Max=
            if ust_max and ust_max > 20000:
                ust_max = None
    if ust is not None and 10 <= ust <= 20000:
        out["ust_v"] = num_or_none(ust)
    if ust_max is not None and 10 <= ust_max <= 20000:
        out["ust_max_v"] = num_or_none(ust_max)

    m = re.search(r"±\s*\(\s*(\d+)\s*\)\s*steps", xfmr, re.I)
    if m:
        out["plus_minus_steps"] = int(m.group(1))
    m = re.search(
        r"\+\s*\(\s*(\d*)\s*\)\s*/\s*[-−]?\s*\(\s*(\d*)\s*\)\s*steps",
        xfmr,
        re.I,
    )
    if m:
        if m.group(1):
            out["plus_steps"] = int(m.group(1))
        if m.group(2):
            out["minus_steps"] = int(m.group(2))
    return out


def extract_mdu(text: str) -> str:
    found = []
    for m in MDU_ASSIGN_RE.finditer(text):
        tok = re.sub(r"\s+", "", m.group(1).upper()).replace("CMA7", "CMA7")
        tok = tok.replace("SHMDL", "SHM-DL").replace("SHMD", "SHM-D") if "SHM" in tok else tok
        if tok == "CMA7" or tok.startswith("CMA"):
            tok = "CMA7"
        if tok not in found:
            found.append(tok)
    if found:
        return ", ".join(found[:3])
    # packing-slip / PO "+ CMA7" already covered. Last resort: unique token
    # not sitting on the RANGE checkbox line.
    tokens = []
    for line in text.splitlines():
        if RANGE_MDU_LINE.search(line) and len(MDU_TOKEN_RE.findall(line)) >= 2:
            continue
        for m in MDU_TOKEN_RE.finditer(line):
            tok = re.sub(r"\s+", "", m.group(1).upper())
            if tok.startswith("CMA"):
                tok = "CMA7"
            if tok not in tokens:
                tokens.append(tok)
    if len(tokens) == 1:
        return tokens[0]
    return ""


def extract_qty_tai(text: str) -> int | None:
    m = re.search(r"(\d+)\s*台", text)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 80:
            return n
    m = re.search(r"Quantity[:\s]+(\d+)", text, re.I)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 80:
            return n
    m = re.search(r"Количество:\s*(\d+)", text)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 80:
            return n
    return None


def looks_like_po(text: str) -> bool:
    return bool(
        re.search(
            r"(Purchase Order|PURCHASE ORDER|Proforma purchase order|"
            r"ORDINE D.ACQUISTO|PO Number|SALES CONTRACT|Mercado Eletrônico)",
            text,
            re.I,
        )
    )


def empty_row(file: str, folder: str, serial: str, customer: str, family_folder: str) -> dict:
    row = {c: "" for c in SALES_COLS}
    row.update(
        {
            "file": file,
            "folder": folder,
            "serial": serial,
            "customer": customer,
            "family_folder": family_folder,
        }
    )
    return row


def extract_pdf_text(path: Path, max_pages: int = MAX_PAGES, time_limit: float = PER_FILE_S) -> tuple[str, int, str]:
    t0 = time.time()
    pages: list[str] = []
    note = ""
    try:
        with pdfplumber.open(str(path)) as pdf:
            n = min(max_pages, len(pdf.pages))
            for i in range(n):
                if time.time() - t0 > time_limit:
                    note = f"page_timeout_at_{i+1}"
                    break
                try:
                    pages.append(pdf.pages[i].extract_text() or "")
                except Exception as e:
                    pages.append("")
                    note = f"page_err_{i+1}:{type(e).__name__}"
    except Exception as e:
        return "", 0, f"open_err:{type(e).__name__}:{e}"
    raw = "\n".join(pages)
    return raw, len(raw), note


def process_file(path_str: str) -> dict:
    path = Path(path_str)
    folder = path.parent.name
    family_folder = FOLDER_ALIAS.get(folder, folder)
    stem = path.stem
    serial, customer = parse_serial_customer(stem)
    row = empty_row(path.name, folder, serial, customer, family_folder)
    raw, nchars, extract_note = extract_pdf_text(path)
    notes = []
    if extract_note:
        notes.append(extract_note)
        if extract_note.startswith("open_err"):
            row["text_status"] = "parse_failed"
            row["notes"] = "; ".join(notes)
            return row
    if nchars < 80:
        row["text_status"] = "scanned_or_empty"
        if extract_note:
            row["notes"] = "; ".join(notes)
        return row

    text = fold_text(raw)
    hits = find_types(text, family_folder)
    best, alts = pick_type(hits, family_folder)
    xfmr = extract_transformer(text)
    mdu = extract_mdu(text)
    qty = extract_qty_tai(raw)  # 台 needs original CJK
    rng = range_unit_count(serial)

    row["text_status"] = "ok"
    if looks_like_po(text) and not DATA_MARKERS.search(text):
        notes.append("po_not_os")

    if best:
        row["sold_type"] = best.get("sold_type") or ""
        row["type_alts"] = alts
        row["phases"] = best.get("phases") or ""
        row["current_a"] = best.get("current_a") if best.get("current_a") is not None else ""
        row["connection"] = best.get("connection") or ""
        row["um_kv"] = best.get("um_kv") if best.get("um_kv") is not None else ""
        row["selector_size"] = best.get("selector_size") or ""
        row["tap_code"] = best.get("tap_code") or ""
        unit = best.get("unit_count") or 1
        if unit == 1:
            unit = qty or rng or 1
        row["unit_count"] = unit
        if best.get("source") == "oltc_data":
            notes.append("type_from_oltc_data")
        elif best.get("source") in {"compact", "fileish"}:
            notes.append("type_from_compact")
        else:
            notes.append(f"type_from_{best.get('source')}")
    else:
        row["unit_count"] = qty or rng or ""
        notes.append("no_type")

    for k in (
        "mva", "rated_kv", "i_a", "i_max_a", "ust_v", "ust_max_v",
        "plus_minus_steps", "plus_steps", "minus_steps",
    ):
        if k in xfmr and xfmr[k] is not None:
            row[k] = xfmr[k]

    row["mdu"] = mdu
    row["notes"] = "; ".join(notes)
    return row


def list_pdfs(root: Path) -> list[Path]:
    files: list[Path] = []
    for d in sorted(root.iterdir()):
        if not d.is_dir():
            continue
        if d.name in SKIP_FOLDERS:
            continue
        if d.name not in INCLUDE_FOLDERS:
            continue
        files.extend(sorted(d.glob("*.pdf")) + sorted(d.glob("*.PDF")))
    # de-dup case on Windows
    seen = set()
    out = []
    for p in files:
        k = str(p).lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def write_xlsx(rows: list[dict], path: Path, summary_rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "sales"
    header_font = Font(bold=True)
    ws.append(SALES_COLS)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)
    numeric = {
        "current_a", "um_kv", "unit_count", "mva", "rated_kv",
        "i_a", "i_max_a", "ust_v", "ust_max_v",
        "plus_minus_steps", "plus_steps", "minus_steps",
    }
    for row in rows:
        ws.append([row.get(c, "") if row.get(c, "") is not None else "" for c in SALES_COLS])
        r = ws.max_row
        for i, c in enumerate(SALES_COLS, 1):
            if c in numeric and row.get(c) not in ("", None):
                ws.cell(r, i).value = row[c]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "file": 42, "folder": 10, "serial": 22, "customer": 28,
        "family_folder": 12, "sold_type": 36, "type_alts": 36,
        "notes": 32,
    }
    for i, c in enumerate(SALES_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    sm = wb.create_sheet("summary")
    sm_cols = [
        "folder", "files", "with_sold_type", "with_i", "with_mva",
        "scanned_or_empty", "parse_failed", "type_rate",
    ]
    sm.append(sm_cols)
    for cell in sm[1]:
        cell.font = header_font
    for r in summary_rows:
        sm.append([r.get(c, "") for c in sm_cols])
    sm.freeze_panes = "A2"
    for i, c in enumerate(sm_cols, 1):
        sm.column_dimensions[get_column_letter(i)].width = 18
    wb.save(path)


def build_summary(rows: list[dict]) -> list[dict]:
    by = defaultdict(list)
    for r in rows:
        by[r.get("folder") or "?"].append(r)

    def pack(name: str, rs: list[dict]) -> dict:
        n = len(rs)
        with_type = sum(1 for x in rs if x.get("sold_type"))
        with_i = sum(1 for x in rs if x.get("i_a") not in ("", None) or x.get("i_max_a") not in ("", None))
        with_mva = sum(1 for x in rs if x.get("mva") not in ("", None))
        scanned = sum(1 for x in rs if x.get("text_status") == "scanned_or_empty")
        failed = sum(1 for x in rs if x.get("text_status") == "parse_failed")
        return {
            "folder": name,
            "files": n,
            "with_sold_type": with_type,
            "with_i": with_i,
            "with_mva": with_mva,
            "scanned_or_empty": scanned,
            "parse_failed": failed,
            "type_rate": round(with_type / n, 3) if n else 0,
        }

    out = [pack(k, by[k]) for k in sorted(by)]
    out.append(pack("TOTAL", rows))
    return out


def print_verify(rows: list[dict]) -> None:
    n = len(rows)
    with_type = [r for r in rows if r.get("sold_type")]
    with_i = [
        r for r in rows
        if r.get("i_a") not in ("", None) or r.get("i_max_a") not in ("", None)
    ]
    with_mva = [r for r in rows if r.get("mva") not in ("", None)]
    print(f"total rows: {n}")
    print(f"rows with sold_type: {len(with_type)} ({len(with_type)/n:.1%})" if n else "rows with sold_type: 0")
    print(f"rows with i_max_a or i_a: {len(with_i)}")
    print(f"rows with mva: {len(with_mva)}")
    print("top 15 sold_type:")
    for t, c in Counter(r["sold_type"] for r in with_type).most_common(15):
        print(f"  {c:4d}  {t}")
    print("5 example rows:")
    shown = 0
    for r in with_type:
        print(
            f"  {r.get('serial')} | {r.get('sold_type')} | mva={r.get('mva')} | "
            f"i={r.get('i_a') or r.get('i_max_a')} | um={r.get('um_kv')}"
        )
        shown += 1
        if shown >= 5:
            break
    scanned = sum(1 for r in rows if r.get("text_status") == "scanned_or_empty")
    failed = sum(1 for r in rows if r.get("text_status") == "parse_failed")
    print(f"scanned_or_empty: {scanned}")
    print(f"parse_failed: {failed}")


SELF_TESTS = [
    ("CM III 600 Y 72,5 C 14253W", "CMIII-600Y/72.5C-14253W"),
    ("CMD III 1000 Y 72.5 C 10193W", "CMDIII-1000Y/72.5C-10193W"),
    ("1x CV III 350 D 40.5 12233G", "CVIII-350D/40.5-12233G"),
    ("CM2III-500Y/72.5C-12233G", "CM2III-500Y/72.5C-12233G"),
    ("3xCZI-500/40.5-7+CMA7", "3xCZI-500/40.5-7"),
    ("OLTC HUAMING SHZV III 1000 Y-126/D 10.19.3W", "SHZVIII-1000Y/126D-10193W"),
    ("OLTC Oil Huaming HWV III 400 Y 40.5 kV 18353W + CMA7", "HWVIII-400Y/40.5-18353W"),
    ("MODEL: CV2 III 350 D 145-10193W", "CV2III-350D/145-10193W"),
    ("COM.CM2 III 500 Y 170/B 14273G", "CM2III-500Y/170B-14273G"),
    ("WSL IV 600 D 72.5 6x5(B)", "WSLIV-600D/72.5-6x5B"),
    ("HWV 3 400 Y 17.5 18353W", "HWVIII-400Y/17.5-18353W"),
    ("SYXZZ-40.5-200-9", "SYXZZ-40.5-200-9"),
    ("CV2III-350D/40.5-10193W", "CV2III-350D/40.5-10193W"),
]


def self_test() -> int:
    failed = 0
    for blob, expect in SELF_TESTS:
        hits = find_types(fold_text(blob), "")
        best, _ = pick_type(hits, "")
        got = best.get("sold_type") if best else None
        ok = got == expect
        print(("OK  " if ok else "FAIL"), repr(blob), "→", got, "" if ok else f"(want {expect})")
        if not ok:
            failed += 1
    # transformer snippets
    xf = extract_transformer(fold_text(
        "Rated capacity (105)MVA , decreasing\n"
        "(регулируемое напр-е / regulated voltage 41) kV\n"
        "± (11) steps ступеней\n"
        "I=( ) A Imax.=(533,5) A\n"
        "Ust max.= (1976,7)V Ust min.= ( )V\n"
    ))
    expect_xf = {"mva": 105, "rated_kv": 41, "plus_minus_steps": 11, "i_max_a": 533.5, "ust_max_v": 1976.7}
    for k, v in expect_xf.items():
        got = xf.get(k)
        ok = got == v
        print(("OK  " if ok else "FAIL"), f"xfmr.{k}", got, "" if ok else f"(want {v})")
        if not ok:
            failed += 1
    ser, cust = parse_serial_customer("E-CM2260001-002 意大利TIRONI")
    print(("OK  " if ser == "E-CM2260001-002" else "FAIL"), "serial", ser, cust)
    if ser != "E-CM2260001-002":
        failed += 1
    print(("OK  " if range_unit_count("E-M260053-072") == 20 else "FAIL"),
          "range", range_unit_count("E-M260053-072"))
    if range_unit_count("E-M260053-072") != 20:
        failed += 1
    return failed


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=Path, default=PDF_ROOT)
    ap.add_argument("--out", type=Path, default=OUT_XLSX)
    ap.add_argument("--copy", type=Path, default=OUT_COPY)
    ap.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--self-test", action="store_true")
    ap.add_argument("--probe", type=str, default="")
    args = ap.parse_args(argv)

    if args.self_test:
        nfail = self_test()
        print(f"self-test failures: {nfail}")
        return 1 if nfail else 0

    if args.probe:
        p = Path(args.probe)
        row = process_file(str(p))
        for k in SALES_COLS:
            print(f"{k:20s} {row.get(k)!r}")
        return 0

    files = list_pdfs(args.root)
    if args.limit:
        files = files[: args.limit]
    print(f"pdfs: {len(files)} workers={args.workers}", flush=True)
    if not files:
        print("no pdfs", file=sys.stderr)
        return 1

    rows: list[dict] = []
    t0 = time.time()
    if args.workers <= 1:
        for i, p in enumerate(files, 1):
            try:
                rows.append(process_file(str(p)))
            except Exception as e:
                r = empty_row(p.name, p.parent.name, p.stem, "", FOLDER_ALIAS.get(p.parent.name, p.parent.name))
                r["text_status"] = "parse_failed"
                r["notes"] = f"{type(e).__name__}:{e}"
                rows.append(r)
            if i % 50 == 0 or i == len(files):
                print(f"  {i}/{len(files)}  {time.time()-t0:.0f}s", flush=True)
    else:
        with ProcessPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(process_file, str(p)): p for p in files}
            done = 0
            for fut in as_completed(futs):
                p = futs[fut]
                done += 1
                try:
                    rows.append(fut.result(timeout=PER_FILE_S + 15))
                except Exception as e:
                    r = empty_row(
                        p.name, p.parent.name, p.stem, "",
                        FOLDER_ALIAS.get(p.parent.name, p.parent.name),
                    )
                    r["text_status"] = "parse_failed"
                    r["notes"] = f"{type(e).__name__}:{e}"
                    rows.append(r)
                if done % 50 == 0 or done == len(files):
                    nt = sum(1 for x in rows if x.get("sold_type"))
                    print(
                        f"  {done}/{len(files)}  typed={nt}  {time.time()-t0:.0f}s",
                        flush=True,
                    )

    # stable order: folder then serial then file
    rows.sort(key=lambda r: (r.get("folder") or "", r.get("serial") or "", r.get("file") or ""))
    summary = build_summary(rows)
    write_xlsx(rows, args.out, summary)
    try:
        write_xlsx(rows, args.copy, summary)
    except Exception as e:
        print(f"desktop copy failed: {e}", flush=True)
    print(f"wrote {args.out}")
    print(f"wrote {args.copy}")
    print_verify(rows)
    print(f"elapsed {time.time()-t0:.0f}s")
    n = len(rows) or 1
    rate = sum(1 for r in rows if r.get("sold_type")) / n
    return 0 if rate >= 0.40 else 2


if __name__ == "__main__":
    raise SystemExit(main())
