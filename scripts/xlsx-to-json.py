#!/usr/bin/env python3
"""Dump docs/2026-os-sales.xlsx sheet `sales` to docs/2026-os-sales.json."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "docs" / "2026-os-sales.xlsx"
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "docs" / "2026-os-sales.json"


def cell(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def main() -> None:
    if not SRC.exists():
        sys.exit(f"missing {SRC}")
    wb = load_workbook(SRC, read_only=True, data_only=True)
    name = "sales" if "sales" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else f"c{i}" for i, h in enumerate(next(rows))]
    sales = []
    for raw in rows:
        rec = {header[i]: cell(raw[i] if i < len(raw) else None) for i in range(len(header))}
        if not any(v not in (None, "") for v in rec.values()):
            continue
        sales.append(rec)
    OUT.write_text(json.dumps({"sales": sales}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(sales)} rows)")


if __name__ == "__main__":
    main()
