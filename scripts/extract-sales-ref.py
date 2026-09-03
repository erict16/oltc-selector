#!/usr/bin/env python3
"""Dump Year=2025 rows from HM reference list to docs/sales-2025-ref.json."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SRC = Path(
    r"C:\Users\TYM\OneDrive - 上海华明电力设备制造有限公司"
    r"\Attachments\Excel\Sales\HM reference list -2019-2025.xlsx"
)
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "docs" / "sales-2025-ref.json"
YEAR = 2025


def main() -> None:
    if not SRC.exists():
        sys.exit(f"missing {SRC}")
    wb = load_workbook(SRC, read_only=True, data_only=True)
    name = next((s for s in wb.sheetnames if s.strip().startswith("2019-2025")), wb.sheetnames[0])
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    # Row 1 empty, row 2 title, row 3 header
    next(rows)
    next(rows)
    header = [str(h).strip() if h is not None else f"c{i}" for i, h in enumerate(next(rows))]
    year_i = next((i for i, h in enumerate(header) if h.lower() == "year"), 1)
    type_i = next((i for i, h in enumerate(header) if "TAP CHANGER TYPE" in h.upper()), 2)
    qty_i = next((i for i, h in enumerate(header) if h.upper() in ("QTY", "QTY.")), 3)
    out = []
    for raw in rows:
        year = raw[year_i] if year_i < len(raw) else None
        try:
            y = int(year)
        except (TypeError, ValueError):
            continue
        if y != YEAR:
            continue
        typ = raw[type_i] if type_i < len(raw) else None
        if not typ or not str(typ).strip():
            continue
        qty = raw[qty_i] if qty_i < len(raw) else 1
        try:
            q = int(qty) if qty not in (None, "") else 1
        except (TypeError, ValueError):
            q = 1
        out.append({"year": y, "sold_type": str(typ).strip(), "qty": q})
    wb.close()
    OUT.write_text(json.dumps({"year": YEAR, "source": SRC.name, "rows": out}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(out)} rows, year={YEAR})")


if __name__ == "__main__":
    main()
